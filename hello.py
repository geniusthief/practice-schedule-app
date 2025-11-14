import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import tempfile
from pulp import LpProblem, LpMaximize, LpVariable, lpSum, LpBinary, LpInteger, LpStatus
import string

# --- ページ設定 ---
st.set_page_config(page_title="卓球部練習スケジュール最適化", layout="wide")
st.title("🏓 卓球部 練習シフト最適化ツール")

# --- Excelアップロード ---
uploaded_file = st.file_uploader("📂 Excelファイルをアップロードしてください", type=["xlsx"])
if uploaded_file is None:
    st.info("👆 Excelファイルをアップロードしてください。")
    st.stop()

# --- Workbook 読み込み ---
tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
tmpf.write(uploaded_file.read())
tmpf.flush()
book = load_workbook(tmpf.name)

# --- r_timeシート表示・編集 ---
st.subheader("🗓️ 可用性（r_time）")
r_time_df = pd.read_excel(tmpf.name, sheet_name="r_time")
edited_r_time = st.data_editor(r_time_df, num_rows="dynamic", key="r_time_edit")

# # --- day_limitsシート表示・編集 ---
# st.subheader("⚙️ 曜日ごとの人数制約（day_limits）")
# day_limits_df = pd.read_excel(tmpf.name, sheet_name="day_limits")
# edited_day_limits = st.data_editor(day_limits_df, num_rows="dynamic", key="day_limits_edit")

# --- チア日選択 ---
st.subheader("🎽 チアの日設定")
cheer_days = st.multiselect("チアのある曜日を選択", ["火", "水", "木", "金"], default=["火", "金"])

# --- 重み設定（任意） ---
with st.sidebar:
    st.header("重みパラメータ")
    w1 = st.number_input("授業直後スコア (w1)", value=100.0)
    w2 = st.number_input("連続練習スコア (w2)", value=0.5)
    w3 = st.number_input("人数スコア (w3)", value=1.0)

# --- 最適化関数 ---
def run_optimization_from_workbook(book, cheer_days, w1, w2, w3):
    sheet_rt = book['r_time']
    sheet_len = book['w_len']
    sheet_day = book['day_limits']

    # 部員数を自動で取得
    num_members = 0
    for i in range(1, 100):
        if sheet_rt.cell(row=i + 1, column=1).value is None:
            break
        num_members += 1

    I = [i + 1 for i in range(num_members)]
    T = [i + 1 for i in range(8)]   # 時間帯 13〜20時 → 1〜8
    D = [i + 1 for i in range(4)]   # 曜日 火〜金 → 1〜4
    L = [1, 2, 3]
    L_s = {s: [l for l in L if s + l - 1 <= 8] for s in T}
    labels = list(string.ascii_uppercase)

    # r_time マップ
    r_map = {2: 1, 3: 3, 4: 5, 5: 7}
    r_time = {}
    for i in I:
        for d in D:
            val = sheet_rt.cell(row=i + 1, column=d + 1).value
            if val in r_map:
                r_time[i, d] = r_map[val]
            else:
                r_time[i, d] = None

    # availability の自動生成
    a = {}
    for i in I:
        for d in D:
            start = r_time[i, d]
            for t in T:
                a[i, t, d] = 0
            if start:
                for t in range(start, 9):
                    a[i, t, d] = 1

    # w_len 読み込み
    w_len = {l: sheet_len.cell(row=l + 1, column=2).value for l in L}

    # --- day_chia の作成（day_limits シート不要） ---
    weekday_map = {1: '火', 2: '水', 3: '木', 4: '金'}
    day_chia = {}
    for d in range(1, 5):
        day_chia[d] = weekday_map[d] in cheer_days

    # min/max
    day_min = {}
    day_max = {}
    for d in range(1, 5):
        if day_chia[d]:
            day_min[d] = 2
            day_max[d] = 9
        else:
            day_min[d] = 2
            day_max[d] = 18

    # ideal / w_num
    ideal = {}
    w_num = {d: {} for d in D}
    N_range = list(range(3, num_members + 1))
    for d in D:
        if day_chia[d] is not None:
            ideal[d] = 8
            for n in N_range:
                w_num[d][n] = max(0.0, 1.0 - 0.1 * abs(n - ideal[d]))
        else:
            ideal[d] = None
            for n in N_range:
                w_num[d][n] = 1.0
            
    # 問題定義（最大化）
    prob = LpProblem("practice_schedule", LpMaximize)

    forbidden_start = [2, 4, 6, 8]  # 14時,16時,18時,20時

    # 変数
    x = {(i, t, d): LpVariable(f"x_{i}_{t}_{d}", cat=LpBinary) for i in I for t in T for d in D}
    y = {(t, d): LpVariable(f"y_{t}_{d}", cat=LpBinary) for t in T for d in D}
    z = {}
    for i in I:
        for d in D:
            for s in T:
                if s in forbidden_start:
                    continue
                for l in L_s.get(s, []):
                    z[(i, s, d, l)] = LpVariable(f"z_{i}_{s}_{d}_{l}", cat=LpBinary)
    v = {(t, d, n): LpVariable(f"v_{t}_{d}_{n}", cat=LpBinary) for t in T for d in D for n in N_range}
    num_td = {(t, d): LpVariable(f"num_{t}_{d}", lowBound=0, cat=LpInteger) for t in T for d in D}

    # 制約: x <= a
    for i in I:
        for t in T:
            for d in D:
                if a[i, t, d] == 0:
                    prob += x[i, t, d] == 0
                else:
                    prob += x[i, t, d] <= 1

    # 週3回以上
    for i in I:
        prob += lpSum([x[i, t, d] for t in T for d in D]) >= 3

    # 一日の最大3時間
    for i in I:
        for d in D:
            prob += lpSum([x[i, t, d] for t in T]) <= 3

    # 人数制約
    for t in T:
        for d in D:
            available = sum(a[i, t, d] for i in I)
            if available < day_min[d]:
                for i in I:
                    prob += x[i, t, d] == 0
                continue
            prob += num_td[(t, d)] == lpSum([x[i, t, d] for i in I])
            prob += num_td[(t, d)] >= day_min[d]
            prob += num_td[(t, d)] <= day_max[d]

    # z 一意
    for i in I:
        for d in D:
            prob += lpSum([z[(i, s, d, l)] for s in T if s not in forbidden_start for l in L_s.get(s, []) if (i, s, d, l) in z]) <= 1

    # z->x
    for i in I:
        for d in D:
            for t in T:
                prob += x[i, t, d] == lpSum(
                    [z[(i, s, d, l)] for s in T if s not in forbidden_start for l in L_s.get(s, []) if (i, s, d, l) in z and s <= t < s + l]
                )

    # 飛び飛び禁止
    for i in I:
        for d in D:
            if len(T) >= 3:
                for t in T[1:-1]:
                    prob += x[i, t, d] <= x[i, t - 1, d] + x[i, t + 1, d]
            t_first = T[0]
            if len(T) >= 2:
                prob += x[i, t_first, d] <= x[i, t_first + 1, d]
            t_last = T[-1]
            if len(T) >= 2:
                prob += x[i, t_last, d] <= x[i, t_last - 1, d]

    # v の一意性 + 人数に一致（Big-M）
    M = len(I)
    for t in T:
        for d in D:
            prob += lpSum([v[(t, d, n)] for n in N_range]) == 1
            for n in N_range:
                prob += num_td[(t, d)] - n <= (1 - v[(t, d, n)]) * M
                prob += n - num_td[(t, d)] <= (1 - v[(t, d, n)]) * M

    # 目的関数
    term1 = lpSum([x[i, r_time[i, d], d] for i in I for d in D if r_time[i, d] is not None and r_time[i, d] in T])
    term2 = lpSum([ (w_len[l] * z[(i, s, d, l)]) 
                    for i in I for d in D for s in T if s not in forbidden_start for l in L_s.get(s, []) 
                    if (i, s, d, l) in z ])
    term3 = lpSum([ w_num[d][n] * v[(t, d, n)] for t in T for d in D for n in N_range ])

    prob += w1 * term1 + w2 * term2 + w3 * term3

    # solve
    prob.solve()

    result_info = {"status": LpStatus[prob.status]}

    # 結果出力
    def write_result_sheet(x_vars, fallback=False):
        if 'result' in book.sheetnames:
            book.remove(book['result'])
        result_sheet = book.create_sheet('result')
        weekday_map = {1: '火', 2: '水', 3: '木', 4: '金'}
        for d in D:
            cell = result_sheet.cell(row=1, column=1 + d)
            cell.value = f"{weekday_map[d]}曜"
            cell.alignment = Alignment(horizontal='center')
            result_sheet.column_dimensions[cell.column_letter].width = 20
            
        display_rows = [1, 3, 5, 7]
        time_map = {1: "2限", 3: "3限", 5: "4限", 7: "5限"}
        for i, t in enumerate(display_rows, start=2):
            cell = result_sheet.cell(row=i, column=1)
            cell.value = time_map[t]
            cell.alignment = Alignment(horizontal='center')
            result_sheet.column_dimensions['A'].width = 12
        # for t in T:
        #     cell = result_sheet.cell(row=1 + t, column=1)
        #     cell.value = f"{12 + t}時"
        #     cell.alignment = Alignment(horizontal='center')
        #     result_sheet.column_dimensions['A'].width = 12


        # --- 名前リストを取得 ---
        names_list = edited_r_time.iloc[:, 0].tolist()  # 1列目が名前列

        # --- 同じ曜日では一度だけ名前を表示する ---
        for d in D:
            for i in I:
                name = names_list[i - 1]
                # 部員 i がこの曜日 d に出る時間帯（奇数時のみ）
                if fallback:
                    active_times = [t for t in display_rows if a[i, t, d] >= 1]
                else:
                    active_times = [t for t in display_rows if x_vars[(i, t, d)].value() is not None and x_vars[(i, t, d)].value() >= 0.5]

                if not active_times:
                    continue

                # その曜日で最も早い時間にだけ表示
                first_t = min(active_times)
                row = 1 + display_rows.index(first_t) + 1
                col = 1 + d
                cell = result_sheet.cell(row=row, column=col)
                prev = cell.value if cell.value else ''
                names = prev.split(',') if prev else []
                if name not in names:
                    names.append(name)
                cell.value = ",".join(names)
                cell.alignment = Alignment(wrap_text=True, horizontal='center')
                cell.font = Font(size=12)

        # for i in I:
        #     name = names_list[i - 1]  # ← edited_r_time の名前列から取得
        #     for t in display_rows:  # ← 偶数時を除外
        #         for d in D:
        #             if x[(i, t, d)].value() is not None and x[(i, t, d)].value() >= 0.5:
        #                 row = 1 + display_rows.index(t) + 1  # 2限がrow=2, 3限がrow=3 ...
        #                 col = 1 + d
        #                 cell = result_sheet.cell(row=row, column=col)
        #                 prev = cell.value if cell.value else ''
        #                 names = prev.split(',') if prev else []
        #                 if name not in names:
        #                     names.append(name)
        #                 cell.value = ",".join(names)
        #                 cell.alignment = Alignment(wrap_text=True, horizontal='center')
        #                 cell.font = Font(size=12)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        book.save(tmp.name)
        return tmp.name

        # スコア計算
        score1 = sum(x[(i, r_time[i, d], d)].value() for i in I for d in D if r_time[i, d] is not None)
        score2 = sum((w_len[l] * z[(i, s, d, l)].value()) for i in I for d in D for s in T if s not in forbidden_start for l in L_s.get(s, []) if (i, s, d, l) in z)
        score3 = sum((w_num[d][n] * v[(t, d, n)].value()) for t in T for d in D for n in N_range)
        result_info.update({
            'score1': score1, 'score2': score2, 'score3': score3,
            'weighted1': w1 * score1, 'weighted2': w2 * score2, 'weighted3': w3 * score3,
            'total_score': w1 * score1 + w2 * score2 + w3 * score3
        })

    # --- 最適化成功なら通常出力 ---
    if LpStatus[prob.status] in ("Optimal", "Optimal Solution Found", "Optimal (or near optimal)"):
        result_info['output_path'] = write_result_sheet(x_vars=x)
    else:
        # --- fallback: 可用性通り出力 ---
        result_info['output_path'] = write_result_sheet(x_vars=x, fallback=True)

    return result_info


# --- 最適化実行ボタン ---
run_button = st.button("最適化を実行")

if run_button:
    # --- edited_r_time を安全にクリーニングして sheet に反映（削除・追加どちらでも対応） ---
    clean_df = edited_r_time.copy()

    # 1) 名前列の列名を取得（最左列が名前である前提）
    name_col = clean_df.columns[0]

    # 2) 名前列の余分な空白除去・文字列化
    clean_df[name_col] = clean_df[name_col].astype(str).str.strip()

    # 3) 空文字 / 'nan' / 'None' に相当する行を削除
    mask_valid = (~clean_df[name_col].isna()) & (clean_df[name_col] != "") & (clean_df[name_col].str.lower() != "nan") & (clean_df[name_col].str.lower() != "none")
    clean_df = clean_df[mask_valid].reset_index(drop=True)

    for col in clean_df.columns[1:]:
        clean_df[col] = pd.to_numeric(clean_df[col], errors='coerce')

    # 4) その他の列も NaN を None に（openpyxl 書き込み時の扱いを安定させる）
    clean_df = clean_df.where(pd.notnull(clean_df), None)

    # 5) 元の r_time シートを消して再作成（完全上書き）
    if "r_time" in book.sheetnames:
        book.remove(book["r_time"])
    sheet_rt = book.create_sheet("r_time")

    # 6) ヘッダ行を書き込む（DataFrame のカラム名をそのまま）
    for j, col_name in enumerate(clean_df.columns, start=1):
        sheet_rt.cell(row=1, column=j, value=str(col_name))

    # 7) データ本体を書き込み（2行目以降）
    for i, row in enumerate(clean_df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            # None はそのまま None（空セル）
            sheet_rt.cell(row=i, column=j, value=val)

    # ✅ 8) 変更を確実に反映させるため、一時保存＆再読み込み
    import tempfile
    tmp_rewrite = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    book.save(tmp_rewrite.name)
    book = load_workbook(tmp_rewrite.name)

    with st.spinner('最適化モデルを作成・解いています...（数秒〜数分かかる場合があります）'):
        info = run_optimization_from_workbook(book, cheer_days, w1, w2, w3)

    st.subheader('最適化結果')
    st.write('モデルステータス:', info.get('status'))
    if info.get('output_path'):
        df = pd.read_excel(info['output_path'], sheet_name='result', index_col=None)
        st.subheader('割当表 (result シート)')
        st.dataframe(df)
        with open(info['output_path'], 'rb') as f:
            data = f.read()
        st.download_button('結果（practice_result.xlsx）をダウンロード', data, file_name='practice_result.xlsx')

        # 最適化成功ならスコアを表示
        if info.get('status') in ("Optimal", "Optimal Solution Found", "Optimal (or near optimal)"):
            st.metric('合計スコア', f"{info.get('total_score'):.2f}")
            st.write('目的関数内訳:')
            st.write(f"授業直後スコア: {info.get('weighted1'):.2f}")
            st.write(f"連続練習スコア: {info.get('weighted2'):.2f}")
            st.write(f"人数スコア: {info.get('weighted3'):.2f}")
    else:
        st.error('実行可能な解が見つかりませんでした。')
else:
    st.info('準備ができたら「最適化を実行」ボタンを押してください。')


